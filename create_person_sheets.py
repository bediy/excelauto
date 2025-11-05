import re
from pathlib import Path

import openpyxl

# 人员信息文件与目标工作簿路径
INFO_FILE = Path(r"D:\Test\2025年驾校考核人员信息汇总.xlsx")
TARGET_FILE = Path(r"D:\Test\工作簿.xlsx")
TEMPLATE_SHEET_NAME = "肖龙飞"


def extract_digits(value: object) -> str:
    """提取字符串中的数字部分，用于获得纯数字士兵证号。"""
    if value is None:
        return ""
    text = str(value).strip()
    return "".join(re.findall(r"\d", text))


def main() -> None:
    """读取人员资料并基于模板生成或更新个人工作表。"""
    # 读取人员信息工作簿
    info_wb = openpyxl.load_workbook(INFO_FILE, data_only=True)
    info_ws = info_wb["Sheet1"]

    # 从第3行开始逐行读取姓名、士兵证号、身份证号
    entries: list[tuple[str, str, str]] = []
    for row in info_ws.iter_rows(min_row=3, min_col=3, max_col=5, values_only=True):
        name, soldier_id, id_card = row

        if not name:
            continue

        name_str = str(name).strip()
        if not name_str:
            continue

        soldier_digits = extract_digits(soldier_id)
        id_card_str = str(id_card).strip() if id_card else ""
        entries.append((name_str, soldier_digits, id_card_str))

    info_wb.close()

    # 打开目标工作簿，基于模板复制新表并填写数据
    target_wb = openpyxl.load_workbook(TARGET_FILE)
    if TEMPLATE_SHEET_NAME not in target_wb.sheetnames:
        raise ValueError(f"模板工作表“{TEMPLATE_SHEET_NAME}”不存在")

    template_ws = target_wb[TEMPLATE_SHEET_NAME]
    existing_names = set(target_wb.sheetnames)

    for name, soldier_id, id_card in entries:
        # 如果工作表不存在则复制模板，否则直接覆盖B3/D3/B4
        if name not in existing_names:
            ws = target_wb.copy_worksheet(template_ws)
            ws.title = name
            existing_names.add(name)
        else:
            ws = target_wb[name]

        # 写入对应字段，保持原有单元格格式
        ws["B3"].value = name
        ws["D3"].value = soldier_id
        ws["B4"].value = id_card

    target_wb.save(TARGET_FILE)
    target_wb.close()

    print(f"已处理 {len(entries)} 位人员信息。")


if __name__ == "__main__":
    main()
